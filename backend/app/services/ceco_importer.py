"""Importador masivo de jerarquía Área → Especialidad → CC desde Excel.

Formato esperado (columnas, no importa el orden — buscadas por nombre):
    Cod01, Area, Cod02, Especialidad, Cod03, CentroCosto,
    TipoCosto (opcional), CodigoCeco (opcional)

Idempotente:
    - Upsert por (proyecto_id, codarea) para áreas
    - Upsert por (area_id, codespecialidad) para especialidades
    - Upsert por (especialidad_id, codcentrocosto) para CC
    - Todo en una sola transacción; si algo falla, rollback completo.

Retorna un resumen con contadores por nivel para feedback al usuario.
"""
from typing import Any, Dict, List, Optional
from uuid import UUID

import openpyxl
from fastapi import HTTPException, status

from app.database import get_db

# ─── Nombres de columna aceptados (case-insensitive, con aliases) ──────────
COL_ALIASES = {
    "cod01":        ["cod01", "codigo_area", "cod_area", "codarea", "cod_1"],
    "area":         ["area", "nombre_area", "nbrarea", "area_nombre"],
    "cod02":        ["cod02", "codigo_esp", "cod_especialidad", "codespecialidad", "cod_2"],
    "especialidad": ["especialidad", "nombre_esp", "nbrespecialidad", "especialidad_nombre"],
    "cod03":        ["cod03", "codigo_cc", "cod_centrocosto", "codcentrocosto", "cod_3"],
    "cc":           ["centrocosto", "centro_costo", "nombre_cc", "nbrcentrocosto", "cc"],
    "tipocosto":    ["tipocosto", "tipo_costo", "tipo"],
    "codigoceco":   ["codigoceco", "codigo_ceco", "cec_final", "ceco"],
    "descripcion":  ["descripcion", "descentrocosto", "desc"],
}


def _norm(s: Any) -> str:
    """Normaliza un header para matcheo (lower, sin espacios, sin puntuación)."""
    if s is None:
        return ""
    return "".join(c for c in str(s).lower() if c.isalnum() or c == "_")


def _build_col_map(header_row: tuple) -> Dict[str, int]:
    """Mapea nombre-canónico → índice de columna. Case/space insensitive."""
    normalized = [_norm(h) for h in header_row]
    col_map: Dict[str, int] = {}
    for canonical, aliases in COL_ALIASES.items():
        for alias in aliases:
            if alias in normalized:
                col_map[canonical] = normalized.index(alias)
                break
    return col_map


def _get(row: tuple, col_map: Dict[str, int], key: str) -> Any:
    idx = col_map.get(key)
    if idx is None:
        return None
    if idx >= len(row):
        return None
    v = row[idx]
    return v


def _str_or_none(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _int_or_none(v: Any) -> Optional[int]:
    if v is None or v == "":
        return None
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return None


def parse_ceco_workbook(file_bytes: bytes) -> Dict[str, Any]:
    """Parsea el .xlsx en memoria. Retorna:
        {
          "columns_found": {...},
          "rows": [{cod01, area, cod02, esp, cod03, cc, ...}],
          "warnings": [str]
        }
    Levanta HTTPException(400) si falta alguna columna obligatoria.
    """
    from io import BytesIO
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as e:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            f"No se pudo abrir el Excel: {e}"
        )

    # Usa la primera hoja o busca una llamada "CecoGrecia" / "cecos" / "ceco"
    ws = None
    preferred = {"cecogrecia", "cecos", "ceco", "cecoazoramind"}
    for name in wb.sheetnames:
        if _norm(name) in preferred:
            ws = wb[name]
            break
    if ws is None:
        ws = wb[wb.sheetnames[0]]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "El Excel está vacío")

    col_map = _build_col_map(header_row)

    # ── Validación 1: columnas obligatorias presentes ──────────────────────
    required = ["cod01", "area", "cod03", "cc"]
    missing = [r for r in required if r not in col_map]
    if missing:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            f"Faltan columnas obligatorias en el Excel: {', '.join(missing)}. "
            f"Se encontraron: {[h for h in header_row if h]}"
        )

    # Límites de negocio (parametrizables si hace falta)
    MAX_ROWS = 5000
    MAX_NAME_LEN = 255
    MAX_COD_LEN = 30

    # ── Validación 2: parseo fila-por-fila con errores/warnings ───────────
    errors: list[str] = []   # bloqueantes — abortar
    warnings: list[str] = [] # informativos — se saltea la fila
    parsed_rows = []

    # Para chequeo de consistencia cruzada (mismo código, distintos nombres)
    area_names_by_cod: dict[int, str] = {}
    esp_names_by_key: dict[tuple, str] = {}
    seen_cc_keys: set[tuple] = set()  # (cod01, cod02, cod03) — dup detection

    for i, row in enumerate(rows_iter, start=2):
        if row is None or all(v is None or v == "" for v in row):
            continue  # skip completely empty

        cod01 = _int_or_none(_get(row, col_map, "cod01"))
        area = _str_or_none(_get(row, col_map, "area"))
        cod02 = _int_or_none(_get(row, col_map, "cod02"))
        esp = _str_or_none(_get(row, col_map, "especialidad"))
        cod03 = _int_or_none(_get(row, col_map, "cod03"))
        cc = _str_or_none(_get(row, col_map, "cc"))
        tipocosto = _str_or_none(_get(row, col_map, "tipocosto"))
        codigoceco = _str_or_none(_get(row, col_map, "codigoceco"))
        descripcion = _str_or_none(_get(row, col_map, "descripcion"))

        # 2.a Obligatorios presentes
        if cod01 is None:
            warnings.append(f"Fila {i}: Cod01 vacío o no numérico — saltada")
            continue
        if not area:
            warnings.append(f"Fila {i}: Area vacía — saltada")
            continue
        if cod03 is None:
            warnings.append(f"Fila {i}: Cod03 vacío o no numérico — saltada")
            continue
        if not cc:
            warnings.append(f"Fila {i}: CentroCosto vacío — saltada")
            continue

        # 2.b Rangos de valores numéricos
        if cod01 < 0 or cod01 > 999999:
            warnings.append(f"Fila {i}: Cod01={cod01} fuera de rango (0–999999) — saltada")
            continue
        if cod03 < 0 or cod03 > 999999:
            warnings.append(f"Fila {i}: Cod03={cod03} fuera de rango (0–999999) — saltada")
            continue

        # 2.c Longitud de strings
        if len(area) > MAX_NAME_LEN:
            warnings.append(f"Fila {i}: Area supera {MAX_NAME_LEN} caracteres — saltada")
            continue
        if len(cc) > MAX_NAME_LEN:
            warnings.append(f"Fila {i}: CentroCosto supera {MAX_NAME_LEN} caracteres — saltada")
            continue
        if esp and len(esp) > MAX_NAME_LEN:
            warnings.append(f"Fila {i}: Especialidad supera {MAX_NAME_LEN} caracteres — saltada")
            continue
        if codigoceco and len(codigoceco) > MAX_COD_LEN:
            warnings.append(f"Fila {i}: CodigoCeco supera {MAX_COD_LEN} caracteres — saltada")
            continue

        # Default especialidad si vino vacía
        if cod02 is None:
            cod02 = 0
        if not esp:
            esp = "GENERAL"

        # 2.d Consistencia cruzada: mismo Cod01 → mismo nombre de Area
        prev_area = area_names_by_cod.get(cod01)
        if prev_area is None:
            area_names_by_cod[cod01] = area
        elif prev_area.strip().lower() != area.strip().lower():
            warnings.append(
                f"Fila {i}: Cod01={cod01} ya se usó con el nombre '{prev_area}'; "
                f"aquí dice '{area}'. Se usará el primero — corrige el Excel para evitar ambigüedad."
            )

        # (Cod01, Cod02) → mismo nombre de Especialidad
        esp_key = (cod01, cod02)
        prev_esp = esp_names_by_key.get(esp_key)
        if prev_esp is None:
            esp_names_by_key[esp_key] = esp
        elif prev_esp.strip().lower() != esp.strip().lower():
            warnings.append(
                f"Fila {i}: Cod01={cod01}, Cod02={cod02} ya se usó como '{prev_esp}'; "
                f"aquí dice '{esp}'. Se usará el primero."
            )

        # 2.e Duplicados exactos de CC dentro del mismo Excel
        cc_key = (cod01, cod02, cod03)
        if cc_key in seen_cc_keys:
            warnings.append(
                f"Fila {i}: CC duplicado (Cod01={cod01}, Cod02={cod02}, Cod03={cod03}). "
                f"Se usará la primera aparición."
            )
            continue
        seen_cc_keys.add(cc_key)

        parsed_rows.append({
            "cod01": cod01,
            "area": area,
            "cod02": cod02,
            "esp": esp,
            "cod03": cod03,
            "cc": cc,
            "tipocosto": tipocosto,
            "codigoceco": codigoceco,
            "descripcion": descripcion,
            "_source_row": i,
        })

        # 2.f Límite duro de filas — abortar antes de saturar
        if len(parsed_rows) > MAX_ROWS:
            errors.append(
                f"El archivo supera el máximo de {MAX_ROWS} filas válidas. "
                f"Divídelo en partes más chicas o contacta a soporte."
            )
            break

    # ── Validación 3: al menos 1 fila válida ──────────────────────────────
    if not parsed_rows and not errors:
        errors.append(
            "Ninguna fila del Excel pasó las validaciones. "
            "Revisa que las columnas obligatorias (Cod01, Area, Cod03, CentroCosto) "
            "tengan valores en todas las filas."
        )

    # Si hay errores bloqueantes, abortar con detalle
    if errors:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "Excel rechazado: " + " | ".join(errors)
        )

    return {
        "columns_found": col_map,
        "sheet_name": ws.title,
        "rows": parsed_rows,
        "warnings": warnings,
    }


def build_template_xlsx() -> bytes:
    """Genera un Excel template con:
    - Header en formato esperado (matchea los aliases del importer)
    - 6 filas de ejemplo que muestran la jerarquía y patrones típicos
    - Columnas dimensionadas para que se lea sin scroll horizontal
    - Header con formato (bold, bg color) para que el usuario no lo edite
    """
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "CecoAzoramind"

    headers = [
        "Cod01", "Area",
        "Cod02", "Especialidad",
        "Cod03", "CentroCosto",
        "TipoCosto", "CodigoCeco", "Descripcion",
    ]
    ws.append(headers)

    # Estilo del header
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="left", vertical="center", wrap_text=False)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
    ws.row_dimensions[1].height = 22

    # Ejemplos: dos áreas, cada una con dos especialidades, cada una con
    # los 4 CC típicos (Mano de Obra / Materiales / Equipos / Subcontratos).
    # Muestra el patrón de códigos que el usuario debe replicar.
    examples = [
        # (cod01, area, cod02, esp, cod03, cc, tipocosto, codigoceco, desc)
        (22, "Trabajos y Obras Preliminares", 11, "Vías de Acceso",         1, "Mano de Obra",          "Costo Directo Inc. IGV", "221101", ""),
        (22, "Trabajos y Obras Preliminares", 11, "Vías de Acceso",         2, "Materiales",            "Costo Directo Inc. IGV", "221102", ""),
        (22, "Trabajos y Obras Preliminares", 11, "Vías de Acceso",         3, "Equipos y Herramientas","Costo Directo Inc. IGV", "221103", ""),
        (22, "Trabajos y Obras Preliminares", 13, "Trabajos Preliminares",  1, "Mano de Obra",          "Costo Directo Inc. IGV", "221301", ""),
        (23, "Obras Iniciales",               10, "Cerco Perimétrico",      1, "Mano de Obra",          "Costo Directo Inc. IGV", "231001", "Ejemplo de descripción"),
        (23, "Obras Iniciales",               10, "Cerco Perimétrico",      2, "Materiales",            "Costo Directo Inc. IGV", "231002", ""),
    ]
    for row in examples:
        ws.append(row)

    # Ancho de columnas
    widths = [7, 32, 7, 30, 7, 26, 22, 12, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze pane debajo del header
    ws.freeze_panes = "A2"

    # Segunda hoja con instrucciones
    ws2 = wb.create_sheet(title="Instrucciones")
    instrucciones = [
        ["Importador de CECOs — Guía rápida"],
        [""],
        ["Columnas OBLIGATORIAS:"],
        ["  Cod01", "Código numérico del área (ej: 22)"],
        ["  Area", "Nombre del área (ej: Trabajos y Obras Preliminares)"],
        ["  Cod03", "Código numérico del centro de costo (ej: 1)"],
        ["  CentroCosto", "Nombre del CC (ej: Mano de Obra)"],
        [""],
        ["Columnas OPCIONALES:"],
        ["  Cod02", "Código numérico de la especialidad (default: 0)"],
        ["  Especialidad", "Nombre de la especialidad (default: GENERAL)"],
        ["  TipoCosto", "ej: Costo Directo Inc. IGV, Indirecto, Gastos Generales"],
        ["  CodigoCeco", "Código concatenado del CC (ej: 221101)"],
        ["  Descripcion", "Detalle libre (opcional)"],
        [""],
        ["Reglas:"],
        ["  · Los códigos se manejan por proyecto. Puedes usar 'Cod01=10' en 2 proyectos"],
        ["    distintos con significados diferentes."],
        ["  · La importación es idempotente: si vuelves a subir el mismo Excel, actualiza"],
        ["    los existentes y no duplica."],
        ["  · Se identifican por códigos: (proyecto, Cod01) para áreas,"],
        ["    (Área, Cod02) para especialidades, (Especialidad, Cod03) para CC."],
        [""],
        ["Máximo del archivo: 5 MB"],
    ]
    for row in instrucciones:
        ws2.append(row)
    ws2.cell(row=1, column=1).font = Font(bold=True, size=14, color="1E40AF")
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 70

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_snapshot_xlsx(proyecto_id: UUID) -> bytes:
    """Genera un Excel con el ESTADO ACTUAL de la jerarquía Área/Especialidad/CC
    de un proyecto, en el MISMO formato que el template — para que el admin
    pueda descargar, editar y re-importar (o guardar como respaldo).

    Cero riesgo de perder configuración: el admin siempre tiene un file
    exportable de lo que hay hoy en la DB, sin depender del Excel original.
    """
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    # Query jerarquía completa en un solo round-trip.
    with get_db() as conn, conn.cursor() as cur:
        cur.execute(
            "SELECT descontratoproyecto, nbrproyecto, codproyecto"
            " FROM construccion.m_proyecto WHERE id = %s;",
            (str(proyecto_id),),
        )
        proy = cur.fetchone()
        if not proy:
            raise HTTPException(status.HTTP_404_NOT_FOUND, "Proyecto no encontrado")

        cur.execute(
            """
            SELECT a.codarea, a.nbrarea,
                   e.codespecialidad, e.nbrespecialidad,
                   c.codcentrocosto, c.nbrcentrocosto,
                   c.tipocentrocosto, c.codigo_ceco, c.descentrocosto
              FROM construccion.m_area a
              LEFT JOIN construccion.m_especialidad e
                     ON e.area_id = a.id AND e.flgactivoespecialidad = true
              LEFT JOIN construccion.m_centrocosto c
                     ON c.especialidad_id = e.id AND c.flgactivocentrocosto = true
             WHERE a.proyecto_id = %s AND a.flgactivoarea = true
             ORDER BY a.codarea::int, e.codespecialidad::int NULLS LAST,
                      c.codcentrocosto::int NULLS LAST;
            """,
            (str(proyecto_id),),
        )
        rows = cur.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "CecoAzoramind"

    headers = [
        "Cod01", "Area",
        "Cod02", "Especialidad",
        "Cod03", "CentroCosto",
        "TipoCosto", "CodigoCeco", "Descripcion",
    ]
    ws.append(headers)

    # Mismo estilo de header que el template (consistencia visual).
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="left", vertical="center", wrap_text=False)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
    ws.row_dimensions[1].height = 22

    for r in rows:
        ws.append([
            r["codarea"],
            r["nbrarea"],
            r["codespecialidad"] or "",
            r["nbrespecialidad"] or "",
            r["codcentrocosto"] or "",
            r["nbrcentrocosto"] or "",
            r["tipocentrocosto"] or "",
            r["codigo_ceco"] or "",
            r["descentrocosto"] or "",
        ])

    widths = [7, 32, 7, 30, 7, 26, 22, 12, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # Metadata en hoja aparte — ayuda a identificar el snapshot al reabrirlo.
    from datetime import datetime
    ws2 = wb.create_sheet(title="Info")
    proy_label = proy["descontratoproyecto"] or proy["nbrproyecto"] or f"Código {proy['codproyecto']}"
    info = [
        ["Snapshot de configuración — Azoramind Tareo"],
        [""],
        ["Proyecto", proy_label],
        ["Código proyecto", proy["codproyecto"]],
        ["Filas exportadas", len(rows)],
        ["Generado", datetime.now().strftime("%Y-%m-%d %H:%M")],
        [""],
        ["Este archivo puede editarse y re-importarse en Configuración → Áreas → Importar Excel."],
        ["La importación es idempotente: actualiza los códigos existentes y agrega los nuevos."],
    ]
    for row in info:
        ws2.append(row)
    ws2.cell(row=1, column=1).font = Font(bold=True, size=14, color="1E40AF")
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 60

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def import_to_proyecto(proyecto_id: UUID, parsed: Dict[str, Any]) -> Dict[str, Any]:
    """Aplica el parseado a un proyecto. Idempotente + transaccional.

    Retorna resumen: {
        areas: {inserted, updated, total},
        especialidades: {inserted, updated, total},
        centros_costo: {inserted, updated, total},
        warnings: [...],
    }
    """
    rows = parsed["rows"]
    warnings = list(parsed["warnings"])

    counters = {
        "areas":         {"inserted": 0, "updated": 0, "total": 0},
        "especialidades":{"inserted": 0, "updated": 0, "total": 0},
        "centros_costo": {"inserted": 0, "updated": 0, "total": 0},
    }

    # Cachés por (cod) → id para evitar N+1
    area_cache: Dict[str, str] = {}       # codarea → id
    esp_cache: Dict[str, str] = {}        # f"{area_id}:{codespecialidad}" → id

    with get_db() as conn, conn.cursor() as cur:
        # Verificar que el proyecto existe
        cur.execute("SELECT id FROM construccion.m_proyecto WHERE id = %s;", (str(proyecto_id),))
        if not cur.fetchone():
            raise HTTPException(status.HTTP_404_NOT_FOUND, "Proyecto no encontrado")

        for r in rows:
            codarea = str(r["cod01"])
            codesp = str(r["cod02"])
            codcc = str(r["cod03"])

            # ── ÁREA ──────────────────────────────────
            area_id = area_cache.get(codarea)
            if area_id is None:
                cur.execute(
                    """
                    INSERT INTO construccion.m_area (proyecto_id, codarea, nbrarea)
                    VALUES (%s, %s, %s)
                    ON CONFLICT (proyecto_id, codarea) DO UPDATE
                        SET nbrarea = EXCLUDED.nbrarea
                    RETURNING id, (xmax = 0) AS was_inserted;
                    """,
                    (str(proyecto_id), codarea, r["area"]),
                )
                row_a = cur.fetchone()
                area_id = row_a["id"]
                area_cache[codarea] = area_id
                if row_a["was_inserted"]:
                    counters["areas"]["inserted"] += 1
                else:
                    counters["areas"]["updated"] += 1
                counters["areas"]["total"] += 1

            # ── ESPECIALIDAD ──────────────────────────
            esp_key = f"{area_id}:{codesp}"
            esp_id = esp_cache.get(esp_key)
            if esp_id is None:
                cur.execute(
                    """
                    INSERT INTO construccion.m_especialidad
                        (area_id, codespecialidad, nbrespecialidad)
                    VALUES (%s, %s, %s)
                    ON CONFLICT (codespecialidad, area_id) DO UPDATE
                        SET nbrespecialidad = EXCLUDED.nbrespecialidad
                    RETURNING id, (xmax = 0) AS was_inserted;
                    """,
                    (area_id, codesp, r["esp"]),
                )
                row_e = cur.fetchone()
                esp_id = row_e["id"]
                esp_cache[esp_key] = esp_id
                if row_e["was_inserted"]:
                    counters["especialidades"]["inserted"] += 1
                else:
                    counters["especialidades"]["updated"] += 1
                counters["especialidades"]["total"] += 1

            # ── CENTRO DE COSTO ───────────────────────
            # Unique key: (especialidad_id, codcentrocosto). No hay unique en
            # DB para este par actualmente — usamos SELECT + INSERT/UPDATE.
            cur.execute(
                """
                SELECT id FROM construccion.m_centrocosto
                 WHERE especialidad_id = %s AND codcentrocosto = %s
                 LIMIT 1;
                """,
                (esp_id, codcc),
            )
            existing = cur.fetchone()
            if existing:
                cur.execute(
                    """
                    UPDATE construccion.m_centrocosto SET
                           nbrcentrocosto = %s,
                           tipocentrocosto = COALESCE(%s, tipocentrocosto),
                           codigo_ceco = COALESCE(%s, codigo_ceco),
                           descentrocosto = COALESCE(%s, descentrocosto)
                     WHERE id = %s;
                    """,
                    (r["cc"], r["tipocosto"], r["codigoceco"],
                     r["descripcion"], existing["id"]),
                )
                counters["centros_costo"]["updated"] += 1
            else:
                cur.execute(
                    """
                    INSERT INTO construccion.m_centrocosto
                        (especialidad_id, codcentrocosto, nbrcentrocosto,
                         tipocentrocosto, codigo_ceco, descentrocosto)
                    VALUES (%s, %s, %s, %s, %s, %s);
                    """,
                    (esp_id, codcc, r["cc"], r["tipocosto"], r["codigoceco"],
                     r["descripcion"]),
                )
                counters["centros_costo"]["inserted"] += 1
            counters["centros_costo"]["total"] += 1

    return {
        **counters,
        "warnings": warnings,
        "processed_rows": len(rows),
        "sheet_name": parsed.get("sheet_name"),
    }
