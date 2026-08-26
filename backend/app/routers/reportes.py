"""Endpoints de reportes: KPIs, dashboard v2 y export Excel formato asistencia."""
from datetime import date, datetime, time, timedelta
from io import BytesIO
from typing import Optional
from uuid import UUID

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.auth.dependencies import get_current_user, require_role
from app.auth.schemas import UserPublic
from app.auth.scoping import get_accessible_proyecto_ids
from app.config import get_settings
from app.database import get_db
from app.repositories import reportes as repo

router = APIRouter(prefix="/api/reportes", tags=["reportes"])


def _default_range() -> tuple[date, date]:
    hoy = date.today()
    return hoy - timedelta(days=30), hoy


@router.get("/kpis")
def kpis(
    desde: date | None = Query(default=None),
    hasta: date | None = Query(default=None),
    _: UserPublic = Depends(get_current_user),
):
    d, h = _default_range()
    d = desde or d
    h = hasta or h
    return {
        "rango": {"desde": d.isoformat(), "hasta": h.isoformat()},
        "generales": repo.kpis_generales(d, h),
        "por_trabajador": repo.horas_por_trabajador(d, h),
        "por_semana": repo.horas_por_semana(d, h),
        "por_centro_costo": repo.horas_por_centro_costo(d, h),
    }


@router.get("/dashboard")
def dashboard(
    desde: date | None = Query(default=None),
    hasta: date | None = Query(default=None),
    proyecto_id: Optional[UUID] = Query(default=None),
    area_id: Optional[UUID] = Query(default=None),
    categoria_id: Optional[UUID] = Query(default=None),
    user: UserPublic = Depends(get_current_user),
):
    """Payload completo para el Dashboard v2.

    Aplica scoping por proyectos accesibles al user (admin bypass = None).
    Un solo round-trip con KPIs + delta período anterior + tendencia diaria
    + top trabajadores + por categoría + por CC + heatmap + alertas + catálogos.
    """
    d, h = _default_range()
    d = desde or d
    h = hasta or h
    scope = get_accessible_proyecto_ids(user)
    if scope is not None and not scope:
        return {
            "kpis": {}, "kpis_prev": {}, "tendencia": [],
            "top_trabajadores": [], "por_categoria": [], "por_cc": [],
            "heatmap": [], "alertas": [],
            "catalogos": {"proyectos": [], "areas": [], "categorias": []},
            "rango": {"desde": d.isoformat(), "hasta": h.isoformat()},
        }
    payload = repo.dashboard_completo(
        d, h,
        proyecto_id=proyecto_id,
        area_id=area_id,
        categoria_id=categoria_id,
        proyecto_ids_scope=scope,
    )
    payload["rango"] = {"desde": d.isoformat(), "hasta": h.isoformat()}
    return payload


def _to_time(v) -> Optional[time]:
    if v is None:
        return None
    if isinstance(v, time):
        return v
    if isinstance(v, datetime):
        return v.time()
    return None


def _minutes_between(a: Optional[time], b: Optional[time]) -> int:
    if not a or not b:
        return 0
    da = datetime.combine(date.today(), a)
    db = datetime.combine(date.today(), b)
    return max(0, int((db - da).total_seconds() // 60))


def _fmt_hm(m: int) -> str:
    if m <= 0:
        return "0:00"
    h, mm = divmod(m, 60)
    return f"{h}:{mm:02d}"


@router.get("/actividades.xlsx")
def export_excel(
    desde: date | None = Query(default=None),
    hasta: date | None = Query(default=None),
    trabajador_id: Optional[UUID] = Query(default=None, description="Opcional: filtra un solo trabajador"),
    user: UserPublic = Depends(require_role("admin", "supervisor")),
) -> StreamingResponse:
    """Reporte "Registro de Control de Asistencia" (formato Grecia neutro).

    Layout:
        Fila 1: [Marca empresa]         [REGISTRO DE CONTROL DE ASISTENCIA]                 [Filtros]
        Fila 3: RAZON SOCIAL | company_name       RUC | company_taxid    FECHA | rango
        Fila 5: N° | Fecha | Apellido y Nombres | DNI | Horario Ingreso |
                Horario Refrigerio Inicio | Horario Refrigerio Fin | Horario Salida |
                Horas Efectivas de Trabajo | Sobretiempo | Firma | Observaciones
        Filas siguientes: registros ordenados por fecha, trabajador.
    """
    settings = get_settings()
    d, h = _default_range()
    d = desde or d
    h = hasta or h

    # Scoping por proyecto (admin ve todo; supervisor solo sus proyectos)
    scope = get_accessible_proyecto_ids(user)
    if scope is not None and not scope:
        rows = []
    else:
        filtro_extra = ""
        params: list = [d, h]
        if trabajador_id:
            filtro_extra += " AND a.trabajador_id = %s"
            params.append(str(trabajador_id))
        if scope is not None:
            filtro_extra += " AND a.proyecto_id = ANY(%s::uuid[])"
            params.append(scope)

        sql = f"""
            SELECT a.fecactividad,
                   t.nbrcompleto        AS apellidos_nombres,
                   COALESCE(t.numidentificacion, '') AS dni,
                   a.horinicio::time    AS horario_ingreso,
                   a.horiniciorefrigerio::time AS refrig_inicio,
                   a.horfinrefrigerio::time    AS refrig_fin,
                   a.horfin::time       AS horario_salida,
                   a.desestadoactividad AS estado,
                   COALESCE(a.desobservaciones, '') AS observaciones
              FROM construccion.m_actividad a
              JOIN construccion.m_trabajador t ON t.id = a.trabajador_id
             WHERE a.fecactividad BETWEEN %s AND %s{filtro_extra}
             ORDER BY a.fecactividad, t.nbrcompleto;
        """
        with get_db() as conn, conn.cursor() as cur:
            cur.execute(sql, tuple(params))
            rows = [dict(r) for r in cur.fetchall()]

    # --- Excel ---
    wb = Workbook()
    ws = wb.active
    ws.title = "CargaMasiva"

    thin = Side(style="thin", color="9CA3AF")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(start_color="F5C542", end_color="F5C542", fill_type="solid")  # dorado Azoramind
    header_font = Font(bold=True, color="1E293B")
    title_font = Font(bold=True, size=14, color="1E40AF")
    label_font = Font(bold=True, color="1E40AF")

    # Fila 1: Marca + Título
    ws.cell(row=1, column=1, value=settings.company_name).font = title_font
    tcell = ws.cell(row=1, column=3, value="REGISTRO DE CONTROL DE ASISTENCIA")
    tcell.font = Font(bold=True, size=13, color="1E293B")
    tcell.alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=10)

    # Fila 3: Razón Social / RUC / Rango
    ws.cell(row=3, column=1, value="RAZÓN SOCIAL").font = label_font
    ws.cell(row=3, column=2, value=settings.company_name)
    ws.cell(row=3, column=5, value="RUC").font = label_font
    ws.cell(row=3, column=6, value=settings.company_taxid or "—")
    ws.cell(row=3, column=9, value="RANGO").font = label_font
    ws.cell(row=3, column=10, value=f"{d.isoformat()} → {h.isoformat()}")

    # Fila 4: agregar día/mes/año del "hasta" como el original (informativo)
    ws.cell(row=4, column=6, value="DÍA:").font = label_font
    ws.cell(row=4, column=7, value=h.day)
    ws.cell(row=4, column=8, value="MES:").font = label_font
    ws.cell(row=4, column=9, value=h.month)
    ws.cell(row=4, column=10, value="AÑO:").font = label_font
    ws.cell(row=4, column=11, value=h.year)

    # Fila 5: Headers de tabla
    headers = [
        "N°", "Fecha", "Apellido y Nombres", "DNI",
        "Horario Ingreso", "Horario Refrigerio Inicio", "Horario Refrigerio Fin",
        "Horario Salida", "Horas Efectivas de Trabajo", "Sobretiempo",
        "Firma", "Observaciones",
    ]
    for col, name in enumerate(headers, start=1):
        c = ws.cell(row=5, column=col, value=name)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border_all
    ws.row_dimensions[5].height = 30

    # Data rows
    jornada_min = int(settings.report_daily_hours * 60)
    for i, r in enumerate(rows, start=6):
        hi = _to_time(r["horario_ingreso"])
        hri = _to_time(r["refrig_inicio"])
        hrf = _to_time(r["refrig_fin"])
        hs = _to_time(r["horario_salida"])

        # Horas efectivas = (salida - ingreso) - refrigerio
        total_min = _minutes_between(hi, hs)
        refrig_min = _minutes_between(hri, hrf) if (hri and hrf) else settings.report_lunch_minutes
        efectivas = max(0, total_min - refrig_min)
        sobretiempo = max(0, efectivas - jornada_min)

        values = [
            i - 5,  # N°
            r["fecactividad"].strftime("%Y-%m-%d") if r["fecactividad"] else "",
            r["apellidos_nombres"],
            r["dni"],
            hi.strftime("%H:%M") if hi else "",
            hri.strftime("%H:%M") if hri else "",
            hrf.strftime("%H:%M") if hrf else "",
            hs.strftime("%H:%M") if hs else "",
            _fmt_hm(efectivas) if hs else "",
            _fmt_hm(sobretiempo) if hs and sobretiempo > 0 else "",
            "",  # Firma (vacía para impresión)
            r["observaciones"],
        ]
        for col, v in enumerate(values, start=1):
            c = ws.cell(row=i, column=col, value=v)
            c.border = border_all
            if col in (1, 5, 6, 7, 8, 9, 10):
                c.alignment = Alignment(horizontal="center")

    # Autoancho
    widths = [5, 12, 32, 12, 10, 12, 12, 10, 12, 12, 18, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A6"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"registro_asistencia_{d.isoformat()}_a_{h.isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
